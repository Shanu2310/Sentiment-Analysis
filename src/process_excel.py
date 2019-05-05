from openpyxl import Workbook, load_workbook
import os


file_name = "/Users/hazique.quazi/myProjects/Sentiment-Analysis/" + "results.xlsx"

def create_or_get_file():
    book = None
    exists = os.path.isfile(file_name)
    if exists:
        book = load_workbook(file_name)
    else:
        book = Workbook()
        book.create_sheet("Training Set")
        book.create_sheet("Result Set")
    
    book.save(file_name)

    return book

def save_data(sheet_name, data_list=[]):
    book = create_or_get_file()
    sheet = book[sheet_name]
    
    for entry in data_list:
        row_data = (entry['text'], entry['sentiment'])
        sheet.append(row_data)
    book.save(file_name)
    book.close()

def del_last_results_file():
    exists = os.path.isfile(file_name)
    if exists:
        os.remove(file_name)    
